# -*- coding:utf-8 -*-
# Time : 2022/3/27 16:07
# Author : Lu
# File : data.py
# Desc :

from config.config import file_path


class ReadWrite:
    def __init__(self):
        # self.file = r'/data/data.txt'
        self.file = ''
        # self.excel = r'D:\Users\lu\PycharmProjects\pythonProject\data\test.xlsx'
        self.excel = file_path
        # self.yaml = r'D:\Users\lu\PycharmProjects\pythonProject\data\test.yml'
        self.yaml = ''

    def txt_read(self):
        list1 = []
        with open(self.file, 'r', encoding='utf-8') as f:
            value = f.readlines()
            f.close()
        for data in value:
            list1.append(data.strip('\n'))
        return list1

    def txt_write(self, username, password):
        f = open(self.file, 'a', encoding='utf-8')
        value = f'{username},{password}\n'
        f.writelines(value)
        f.close()

    def excel_read(self, sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excel)
        table = wb[sheetname]
        rows = table.max_row
        cols = table.max_column
        list_all = []
        for row in range(2, rows + 1):
            list_1 = []
            for col in range(1, cols + 1):
                values = table.cell(row, col).value
                list_1.append(values)
            list_all.append(list_1)
        print(list_all)
        return list_all

    def excel_write(self, sheetname, *args):
        import openpyxl
        wb = openpyxl.load_workbook(self.excel)
        table = wb[sheetname]
        row = table.max_row
        col = len(args)
        for i in range(col):
            table.cell(row + 1, i + 1).value = args[i]
        wb.save(self.excel)

    def mysql_read(self):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='root', database='T31', charset='utf8')
        # 定义游标
        cur = db.cursor()
        sql = 'select * from users where username="Cindy" '
        cur.execute(sql)
        content = cur.fetchall()
        return content

    def mysql_write(self,username,password):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='root', database='T31', charset='utf8')
        # 定义游标
        cur = db.cursor()
        sql = f'insert into users values("{username}","{password}")'
        cur.execute(sql)
        db.commit()

    def yaml_read(self):
        import yaml
        f = open(self.yaml, 'r', encoding='utf-8')
        content = f.read()
        data = yaml.safe_load(content)
        return data

    def yaml_write(self, username, password):
        import yaml
        f = open(self.yaml, 'a', encoding='utf-8')
        data = {'username': username, 'password': password}
        yaml.safe_dump(data, f)
        f.close()
